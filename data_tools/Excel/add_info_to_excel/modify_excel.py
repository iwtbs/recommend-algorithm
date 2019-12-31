#coding:gbk
import sys
import openpyxl

docid_info_dict = {line.strip().split('\t')[0]:line.strip().split('\t')[1] for line in file(sys.argv[1])}

wb = openpyxl.load_workbook(sys.argv[2])
ws = wb.active

ws['M1'] = '【doc附加信息】'.decode("gbk","ignore").encode("utf8","ignore")
ws['M1'].font = openpyxl.styles.Font(bold=True) 

for i in range(2, ws.max_row + 1):
    ws['M' + str(i)] = docid_info_dict.get(ws['B' + str(i)].value, '').decode("gbk","ignore").encode("utf8","ignore")

wb.save(sys.argv[3])

